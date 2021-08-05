from tkinter import *
from tkinter import messagebox
from PIL import Image
from PIL import ImageTk

from threading import Thread
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
import openpyxl as pyxl
import qrcode
import smtplib
import os

from OpenCV import  VideoCapture

#======================================================GUI==============================================================#
class App:
    def __init__(self, video_source=0):
        self.appName = "LADS v2.0"
        self.window = Tk()
        self.window.title(self.appName)
        self.window.resizable(0, 0)
        self.video_source = video_source
        
        #Mainself.menu
        self.menu = Frame(self.window)
        L1 = Label(self.window, text="License Authorization and Driver's Safety", font=15).pack(pady=10)
        self.menu.pack(side = LEFT)
        self.B1 = Button(self.menu, text = "Start", width =15, height =5, command = self.start)
        self.B1.pack()
        self.B2 = Button(self.menu, text="Register", width =15, height =5, command = self.register)
        self.B2.pack()
        self.B2["state"] = "disabled"
        self.B3 = Button(self.menu, text="Recognize again", width =15, height =5, command = self.start)
        self.B3.pack()
        self.B3["state"] = "disabled"
        self.B4 = Button(self.menu, text="Un-authorized\nDriving", width =15, height =5, command = self.unauth_drive)
        self.B4.pack()
        self.B4["state"] = "disabled"
        self.B5 = Button(self.menu, text="Verify through\nQR Code", width =15, height =5, command = self.f_QR)
        self.B5.pack()
        self.B5["state"] = "disabled"
        self.B6 = Button(self.menu, text="Exit", width =15, height =5, command=self.window.destroy)
        self.B6.pack()

        #OpenCV
        self.vid = VideoCapture(self.video_source)
        self.cv = Frame(self.window)
        self.cv.pack(side = LEFT, padx=10)
        self.cv_canvas = Canvas(self.cv, width=int(self.vid.width), height=int(self.vid.height))
        self.cv_canvas.pack()

        #register
        self.f_register = Frame(self.window)
        self.name_field = Entry(self.f_register)
        self.lic_field = Entry(self.f_register)
        self.email_id_field = Entry(self.f_register)
        self.wb = pyxl.load_workbook('admin usage files/License.xlsx')
        self.sheet = self.wb.active
        self.reg_count = 0

        #QrCode
        self.qr = Frame(self.cv)
        self.qr_name = Entry(self.qr)
        self.need_row = []

        #Flags
        self.RECOGNIZE = False
        self.DROWSY = False
        self.reg_flag = False
        self.UNAUTHORIZED_DRIVE = False
        self.QR = False

        self.update()
        self.window.mainloop()

    def update(self):
        #At start
        if not(self.DROWSY or self.RECOGNIZE or self.QR):
            isTrue, frame = self.vid.getFrame()
            if isTrue:
                #Get frames from video source and displays in Tkinter window
                self.photo = ImageTk.PhotoImage(Image.fromarray(frame))
                self.cv_canvas.create_image(0, 0, image=self.photo,anchor = NW)
        #Drowziness
        elif self.DROWSY:
            isTrue, frame = self.vid.ear()
            if isTrue:
                # Get frames from video source and displays in Tkinter window
                self.photo = ImageTk.PhotoImage(Image.fromarray(frame))
                self.cv_canvas.create_image(0, 0, image=self.photo, anchor=NW)
        #QR-Verification
        elif self.QR:
            isTrue, frame, verified = self.vid.getQR(self.need_row[1])
            if isTrue and not verified:
                self.photo = ImageTk.PhotoImage(Image.fromarray(frame))
                self.cv_canvas.create_image(0, 0, image=self.photo, anchor=NW)
            elif verified:
                self.decisions()
            else:
                messagebox.showinfo(title="Data Incorrect",message="Please show the sent QR code")
        #Recognition
        else:
            isTrue, frame, finished = self.vid.recognize()
            if isTrue and not finished:
                self.photo = ImageTk.PhotoImage(Image.fromarray(frame))
                self.cv_canvas.create_image(0, 0, image=self.photo, anchor=NW)
            else:
                self.RECOGNIZE = False
                self.decisions()
        self.window.after(1,self.update)

    #Starts Recognition
    def start(self):
        if not self.cv.winfo_ismapped():
            self.f_register.pack_forget()
            self.qr.pack_forget()
            #self.vid = VideoCapture(0)
            self.cv.pack()
        self.RECOGNIZE = True


    #handles button visibility, Core logic
    def decisions(self):
        #Starting Decision
        if self.vid.authorize:
            self.B1["state"] = "disabled"
            self.B2["state"] = "disabled"
            self.B3["state"] = "disabled"
            self.B4["state"] = "disabled"
            self.B5["state"] = "disabled"
            self.DROWSY = True
            print("authorized")
        if self.reg_flag:
            print("registration finished")
            self.reg_flag = False
            self.vid.authorize = True
            self.B1["state"] = "active"
        elif self.UNAUTHORIZED_DRIVE:
            self.DROWSY = True
        elif self.QR:
            self.QR = False
            self.DROWSY = True
        else:
            if self.DROWSY:
                return
            print("unauthorized")
            self.vid.authorize = True
            self.B1["state"] = "disabled"
            self.B2["state"] = "active"
            self.B3["state"] = "active"
            self.B4["state"] = "active"
            self.B5["state"] = "active"

    #REGISTRATION START
    def register(self):
        if self.cv.winfo_ismapped():
            self.cv.pack_forget()
        self.f_register.pack()
        self.excel()

        # create a Form labels
        heading = Label(self.f_register, text="Registration")
        name = Label(self.f_register, text="Name")
        lic = Label(self.f_register, text="License No.")
        email_id = Label(self.f_register, text="Email id")

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        heading.grid(row=0, column=1)
        name.grid(row=1, column=0)
        lic.grid(row=2, column=0)
        email_id.grid(row=3, column=0)

        # bind method of widget is used for
        # the binding the function with the events
        # whenever the enter key is pressed
        # then call the focus1 function
        self.name_field.bind("<Return>", self.focus1)
        self.lic_field.bind("<Return>", self.focus2)

        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        self.name_field.grid(row=1, column=1, ipadx="100")
        self.lic_field.grid(row=2, column=1, ipadx="100")
        self.email_id_field.grid(row=3, column=1, ipadx="100")


        # call excel function

        # create a Submit Button and place into the root window
        submit = Button(self.f_register, text="Submit", fg="Black",
                        bg="Red", command=self.insert)
        L1 = Label(self.f_register, text = "Please Look at Camera for 3 seconds")
        submit.grid(row=8, column=1)
        L1.grid(row=9, column=1)

    def excel(self):
        # resize the width of columns in
        # excel spreadsheet
        self.sheet.column_dimensions['A'].width = 50
        self.sheet.column_dimensions['B'].width = 50
        self.sheet.column_dimensions['C'].width = 50

        # write given data to an excel spreadsheet
        # at particular location
        self.sheet.cell(row=1, column=1).value = "Name"
        self.sheet.cell(row=1, column=2).value = "License"
        self.sheet.cell(row=1, column=3).value = "Email id"

        # Function to set focus (cursor)

    def focus1(self, event):
        # set focus on the course_field box
        self.lic_field.focus_set()

        # Function to set focus

    def focus2(self, event):
        # set focus on the sem_field box
        self.email_id_field.focus_set()

        # Function for clearing the
        # contents of text entry boxes

    def clear(self):
        # clear the content of text entry box
        self.name_field.delete(0, END)
        self.lic_field.delete(0, END)
        self.email_id_field.delete(0, END)
        self.reg_count = 0

        # Function to take data from GUI
        # window and write to an excel file

    def insert(self):
        # if user not fill any entry
        # then print "empty input"
        if (self.name_field.get() == "" and
                self.lic_field.get() == "" and
                self.email_id_field.get() == ""):

            print("empty input")

        else:
            while self.reg_count <1:
                # assigning the max row and max column
                # value upto which data is written
                # in an excel sheet to the variable
                current_row = self.sheet.max_row
                current_column = self.sheet.max_column

                # get method returns current text
                # as string which we write into
                # excel spreadsheet at particular location
                self.sheet.cell(row=current_row + 1, column=1).value = self.name_field.get()
                self.sheet.cell(row=current_row + 1, column=2).value = self.lic_field.get()
                self.sheet.cell(row=current_row + 1, column=3).value = self.email_id_field.get()

                # save the file
                self.wb.save('admin usage files/License.xlsx')

                # Stor and train images

                # set focus on the name_field box
                self.name_field.focus_set()
                self.reg_flag = True
                self.reg_count += 1
                self.vid.storingImages(self.name_field.get())
            # call the clear() function
            messagebox.showinfo(title="Finished Cpature", message="Finished Capturing Photos, Training now")
            self.vid.trainingImages()
            messagebox.showinfo(title="Training Finished", message= "Finished Training. Have a safe ride!")
            self.clear()
            self.decisions()
    #REGISTRATION END

    #Unauthorized Driving
    def unauth_drive(self):
        print("Unauthorized Driving, taking snapshots")
        now = datetime.now()
        dtString = now.strftime('%H-%M-%S')
        name = "unkown" + dtString
        self.vid.storingImages(name, True)
        self.UNAUTHORIZED_DRIVE = True
        self.decisions()
    #Unauthorized Driving End

    #QR Code
    def f_QR(self):
        if not self.qr.winfo_ismapped():
            self.f_register.pack_forget()
        self.qr.pack(side = BOTTOM)
        heading = Label(self.qr, text="QR Verification")
        name = Label(self.qr, text="Name")
        heading.grid(row = 0, column = 1)
        name.grid(row=1, column = 0)
        self.qr_name.grid(row=1, column=1)
        submit = Button(self.qr, text="Submit", fg="Black", bg="Red", command=self.verifyQr)
        submit.grid(row=2, column =1)

    def verifyQr(self):
        found = False
        for currentrow in range(2,self.sheet.max_row+1):
            if self.sheet["A"+str(currentrow)].value == self.qr_name.get():
                self.need_row.append(self.sheet["A"+str(currentrow)].value)
                self.need_row.append(self.sheet["B" + str(currentrow)].value)
                self.need_row.append(self.sheet["C" + str(currentrow)].value)
                found =True
                break
        if not found:
            messagebox.showinfo(title="Not Found", message="Data Not found. Please Register")
        else:
            qr = qrcode.make(self.need_row[1])
            qr.save("MyQR.png")
            self.sendMail("MyQR.png", self.need_row[2])
            messagebox.showinfo(title="Data Found", message="Please check your mail and scan you QR code with camera")
            self.QR = True

    #send email, parameters: QRcode and mail
    def sendMail(self, imgName, email_send):
        self.imgName = imgName
        self.email_send = email_send
        email_user = 'iamnullnull007@gmail.com'
        img_data = open(self.imgName, 'rb').read()
        msg = MIMEMultipart()
        msg['Subject'] = 'QR CODE FOR AUORTHIZE DRIVER'
        msg['From'] = 'iamnullnull007@gmail.com'
        msg['To'] = 'pvsaikrithik@gmail.com'
        text = MIMEText("QR CODE ")
        msg.attach(text)
        image = MIMEImage(img_data, name=os.path.basename(self.imgName))
        msg.attach(image)

        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.ehlo()
        s.starttls()
        s.ehlo()
        s.login(email_user, 'jeebu1234')
        s.sendmail(email_user, self.email_send, msg.as_string())
        s.quit()
    #QR CODE END


if __name__ == "__main__":
    App()
